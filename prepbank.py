"""
Company question bank.

Two sources feed one list per company:

  * A static bank imported from the ai-resume-agent project - 209 coding
    problems across 10 companies, with worked examples. Shipped as JSON in
    data/ and loaded once at import, because it never changes at runtime.
  * Questions the admin adds through the panel, stored in Firestore. These
    stack onto whichever company they name, creating it if it is new.

prepsets.py stays the owner of the interview-round material (the short Q&A
shown in the preview modal). This module owns the practice-desk questions.
"""
import json
import os
import re
import threading
import time
from urllib.parse import urlparse

_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')

QUESTION_TYPES = ["Coding", "MCQ", "Fill in the blank"]
DIFFICULTIES = ["Easy", "Medium", "Hard"]

# Display names and grouping for the imported companies, carried over from the
# source project so the roster reads the same as where the questions came from.
BANK_COMPANIES = {
    "meta": ("Meta", "FANG"),
    "amazon": ("Amazon", "FANG"),
    "netflix": ("Netflix", "FANG"),
    "google": ("Google", "FANG"),
    "infosys": ("Infosys", "Mass Recruiter"),
    "accenture": ("Accenture", "Mass Recruiter"),
    "capgemini": ("Capgemini", "Mass Recruiter"),
    "ibm": ("IBM", "Mass Recruiter"),
    "github": ("GitHub", "Other"),
    "anthropic": ("Anthropic", "Other"),
}

_lock = threading.Lock()
_custom_cache = None


def _load_json(name):
    path = os.path.join(_DATA_DIR, name)
    try:
        with open(path, encoding='utf-8') as fh:
            return json.load(fh)
    except (OSError, ValueError):
        # A missing or corrupt bank must not take the site down - the admin's
        # own questions still work, and the pages render with what they have.
        return {}


_RAW_BANK = _load_json('company_questions.json')
_EXAMPLES = _load_json('question_examples.json')

_EMPTY_EXAMPLE = {
    "input": "", "output": "",
    "explanation": "Example not yet available for this problem.",
}


def _slug_from_url(url: str) -> str:
    """LeetCode problem slug, which is how the examples file is keyed."""
    path = urlparse(url or "").path.strip('/')
    parts = [p for p in path.split('/') if p]
    return parts[-1] if parts else ""


def _build_bank():
    bank = {}
    for slug, questions in _RAW_BANK.items():
        rows = []
        for index, q in enumerate(questions):
            problem = _slug_from_url(q.get('url', ''))
            rows.append({
                "id": f"bank:{slug}:{index}",
                "source": "bank",
                "type": "Coding",
                "title": q.get('title', 'Untitled'),
                "difficulty": q.get('difficulty', 'Medium'),
                "url": q.get('url', ''),
                "acceptance_pct": q.get('acceptancePct'),
                "frequency_pct": q.get('frequencyPct'),
                "example": _EXAMPLES.get(problem, _EMPTY_EXAMPLE),
                "prompt": "",
                "options": [],
                "answer": "",
            })
        bank[slug] = rows
    return bank


_BANK = _build_bank()


def slugify(name: str) -> str:
    slug = re.sub(r'[^a-z0-9]+', '-', (name or '').strip().lower()).strip('-')
    return slug or 'company'


def bank_slugs() -> list[str]:
    return list(_BANK)


def bank_questions(slug: str) -> list[dict]:
    return _BANK.get(slug, [])


# --- admin-authored questions ----------------------------------------------

def invalidate():
    global _custom_cache
    with _lock:
        _custom_cache = None


def list_custom(db) -> list[dict]:
    """Every admin-added question, newest first. Cached; writes invalidate."""
    global _custom_cache
    if _custom_cache is not None:
        return _custom_cache

    rows = []
    if db is not None:
        try:
            for doc in db.collection('prep_questions').stream():
                row = doc.to_dict() or {}
                row['id'] = doc.id
                row['source'] = 'custom'
                rows.append(row)
        except Exception:
            rows = []

    rows.sort(key=lambda r: r.get('created_at', 0), reverse=True)
    with _lock:
        _custom_cache = rows
    return rows


def add_question(db, data: dict, author: str) -> str:
    company_name = (data.get('company_name') or '').strip()
    slug = slugify(company_name)

    payload = {
        "company_slug": slug,
        "company_name": company_name,
        "type": data.get('type') if data.get('type') in QUESTION_TYPES else 'Coding',
        "difficulty": data.get('difficulty') if data.get('difficulty') in DIFFICULTIES else 'Medium',
        "title": (data.get('title') or '').strip(),
        "prompt": (data.get('prompt') or '').strip(),
        # Only meaningful for MCQ, but stored uniformly so one read serves
        # every type without branching at the call site.
        "options": [o for o in (data.get('options') or []) if o.strip()],
        "answer": (data.get('answer') or '').strip(),
        "explanation": (data.get('explanation') or '').strip(),
        # Starter code shown on the practice desk and loaded into the editor,
        # so a question can hand over a signature rather than a blank file.
        "example_function": (data.get('example_function') or '').rstrip(),
        "url": (data.get('url') or '').strip(),
        "created_by": author,
        "created_at": time.time(),
    }

    ref = db.collection('prep_questions').document()
    ref.set(payload)
    invalidate()
    return ref.id


def delete_question(db, question_id: str) -> None:
    db.collection('prep_questions').document(question_id).delete()
    invalidate()


def custom_for_slug(db, slug: str) -> list[dict]:
    return [q for q in list_custom(db) if q.get('company_slug') == slug]


# --- combined view ----------------------------------------------------------

def questions_for(db, slug: str) -> list[dict]:
    """
    Everything for one company: the imported bank first, then anything the
    admin has added on top. Admin questions come last so a freshly added one
    is not buried mid-list.
    """
    return bank_questions(slug) + custom_for_slug(db, slug)


def company_names(db) -> dict:
    """slug -> display name, across both sources."""
    names = {slug: label for slug, (label, _group) in BANK_COMPANIES.items()}
    for q in list_custom(db):
        slug = q.get('company_slug')
        if slug and slug not in names:
            names[slug] = q.get('company_name') or slug.title()
    return names


def total_questions(db) -> int:
    """Every practice question across every company, both sources."""
    return sum(len(questions_for(db, slug)) for slug in company_names(db))


def counts_for(db, slug: str) -> dict:
    questions = questions_for(db, slug)
    by_difficulty = {d: 0 for d in DIFFICULTIES}
    by_type = {t: 0 for t in QUESTION_TYPES}
    for q in questions:
        by_difficulty[q.get('difficulty', 'Medium')] = \
            by_difficulty.get(q.get('difficulty', 'Medium'), 0) + 1
        by_type[q.get('type', 'Coding')] = by_type.get(q.get('type', 'Coding'), 0) + 1
    return {
        "total": len(questions),
        "by_difficulty": by_difficulty,
        "by_type": by_type,
    }


def build_analytics(db) -> dict:
    """Everything the admin questionnaire dashboard reports."""
    names = company_names(db)

    per_company = []
    totals_difficulty = {d: 0 for d in DIFFICULTIES}
    totals_type = {t: 0 for t in QUESTION_TYPES}
    total = 0
    imported = 0
    authored = 0

    for slug, name in names.items():
        counts = counts_for(db, slug)
        bank_n = len(bank_questions(slug))
        custom_n = len(custom_for_slug(db, slug))
        imported += bank_n
        authored += custom_n
        total += counts['total']
        for key, value in counts['by_difficulty'].items():
            totals_difficulty[key] = totals_difficulty.get(key, 0) + value
        for key, value in counts['by_type'].items():
            totals_type[key] = totals_type.get(key, 0) + value

        per_company.append({
            "slug": slug,
            "name": name,
            "total": counts['total'],
            "imported": bank_n,
            "authored": custom_n,
            "by_difficulty": counts['by_difficulty'],
        })

    per_company.sort(key=lambda c: (-c['total'], c['name']))
    company_count = len(names)

    return {
        "companies": company_count,
        "questions": total,
        "imported": imported,
        "authored": authored,
        # Guarded so an install with no companies reads 0 rather than raising.
        "avg_per_company": round(total / company_count, 1) if company_count else 0,
        "by_difficulty": totals_difficulty,
        "by_type": totals_type,
        # Pre-shaped for the bar macro, which wants label/count rows.
        "by_type_rows": [{"label": k, "count": v} for k, v in totals_type.items()],
        "per_company": per_company,
        "recent": list_custom(db)[:10],
    }


# --- spreadsheet import/export ---------------------------------------------
# Columns are the questionnaire form, in the order it asks for them, so the
# sheet and the form stay recognisably the same thing.
SHEET_COLUMNS = [
    "Company name", "Type of question", "Difficulty", "Short title",
    "Question", "Example function", "Option A", "Option B", "Option C",
    "Option D", "Answer", "Explanation", "Reference link",
]


def build_template_workbook():
    """The blank sheet an admin downloads, fills in and uploads back."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"

    header_fill = PatternFill("solid", fgColor="5B4BDB")
    for col, name in enumerate(SHEET_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[cell.column_letter].width = \
            34 if name in ("Question", "Example function", "Explanation") else 18
    ws.freeze_panes = "A2"

    # Dropdowns rather than free text, so an upload cannot invent a type or a
    # difficulty the app does not understand.
    type_rule = DataValidation(
        type="list", formula1='"{}"'.format(",".join(QUESTION_TYPES)), allow_blank=True)
    diff_rule = DataValidation(
        type="list", formula1='"{}"'.format(",".join(DIFFICULTIES)), allow_blank=True)
    ws.add_data_validation(type_rule)
    ws.add_data_validation(diff_rule)
    type_rule.add("B2:B500")
    diff_rule.add("C2:C500")

    samples = [
        ["Amazon", "Coding", "Easy", "Two Sum",
         "Given an array and a target, return the indices of the two numbers "
         "that add up to the target.",
         "function twoSum(nums, target) {\n    // your code here\n}",
         "", "", "", "", "[0, 1]",
         "Walk the array keeping value -> index in a map.",
         "https://leetcode.com/problems/two-sum"],
        ["Infosys", "MCQ", "Easy", "Binary search complexity",
         "What is the time complexity of binary search on a sorted array?", "",
         "O(1)", "O(log n)", "O(n)", "O(n log n)", "O(log n)",
         "Each comparison halves the remaining search space.", ""],
        ["Wipro", "Fill in the blank", "Medium", "SQL filtering",
         "The ______ clause filters rows after GROUP BY has been applied.", "",
         "", "", "", "", "HAVING",
         "WHERE filters before grouping; HAVING filters after.", ""],
    ]
    for row in samples:
        ws.append(row)
    for row in ws.iter_rows(min_row=2, max_row=1 + len(samples)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    notes = wb.create_sheet("How to use")
    for line in [
        "Fill one question per row on the Questions sheet, then upload the file.",
        "",
        "Company name  required. An existing name adds to that company; a new one creates it.",
        "Type          " + " / ".join(QUESTION_TYPES),
        "Difficulty    " + " / ".join(DIFFICULTIES),
        "Question      required. Supports **bold**, `code`, ```blocks``` and - bullets.",
        "Options       MCQ only, at least two.",
        "Answer        the correct option for MCQ, or the expected result.",
        "",
        "The three sample rows are examples - delete them before uploading, or",
        "they will be imported alongside your own.",
    ]:
        notes.append([line])
    notes.column_dimensions["A"].width = 100
    return wb


def parse_workbook(stream) -> tuple[list[dict], list[str]]:
    """
    Reads an uploaded sheet into question payloads.

    Returns (rows, errors). Rows that fail validation are reported by line
    number and skipped; the valid ones still import, so one bad row does not
    cost the whole upload.
    """
    from openpyxl import load_workbook

    wb = load_workbook(stream, data_only=True, read_only=True)
    ws = wb["Questions"] if "Questions" in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["The sheet is empty."]

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    index = {name.lower(): i for i, name in enumerate(header) if name}

    def cell(row, column):
        i = index.get(column.lower())
        if i is None or i >= len(row) or row[i] is None:
            return ""
        return str(row[i]).strip()

    missing = [c for c in ("Company name", "Question") if c.lower() not in index]
    if missing:
        return [], [f"Missing required column(s): {', '.join(missing)}."]

    parsed, errors = [], []
    for line, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue

        company = cell(row, "Company name")
        prompt = cell(row, "Question")
        if not company or not prompt:
            errors.append(f"Row {line}: needs both a company name and a question.")
            continue

        qtype = cell(row, "Type of question") or "Coding"
        if qtype not in QUESTION_TYPES:
            errors.append(f"Row {line}: '{qtype}' is not a question type.")
            continue

        difficulty = cell(row, "Difficulty") or "Medium"
        if difficulty not in DIFFICULTIES:
            errors.append(f"Row {line}: '{difficulty}' is not a difficulty.")
            continue

        options = [cell(row, f"Option {letter}") for letter in "ABCD"]
        options = [o for o in options if o]
        if qtype == "MCQ" and len(options) < 2:
            errors.append(f"Row {line}: an MCQ needs at least two options.")
            continue

        parsed.append({
            "company_name": company,
            "type": qtype,
            "difficulty": difficulty,
            "title": cell(row, "Short title"),
            "prompt": prompt,
            "example_function": cell(row, "Example function"),
            "options": options,
            "answer": cell(row, "Answer"),
            "explanation": cell(row, "Explanation"),
            "url": cell(row, "Reference link"),
        })

    return parsed, errors


def add_many(db, rows: list[dict], author: str) -> int:
    """Bulk insert, one Firestore batch. Same collection as the form writes."""
    if not rows:
        return 0

    batch = db.batch()
    written = 0
    for data in rows:
        ref = db.collection('prep_questions').document()
        batch.set(ref, {
            "company_slug": slugify(data["company_name"]),
            "company_name": data["company_name"],
            "type": data["type"],
            "difficulty": data["difficulty"],
            "title": data.get("title", ""),
            "prompt": data["prompt"],
            "options": data.get("options", []),
            "answer": data.get("answer", ""),
            "explanation": data.get("explanation", ""),
            "example_function": data.get("example_function", ""),
            "url": data.get("url", ""),
            "created_by": author,
            "created_at": time.time(),
        })
        written += 1
        # Firestore caps a batch at 500 writes.
        if written % 400 == 0:
            batch.commit()
            batch = db.batch()

    batch.commit()
    invalidate()
    return written
